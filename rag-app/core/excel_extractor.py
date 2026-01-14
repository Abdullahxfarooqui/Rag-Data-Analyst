"""
Pure Python Excel extractor - NO pandas, NO numpy, NO segfaults.
Supports both .xlsx and .xls files.
"""
import io
from typing import Dict, Any, List
from pathlib import Path


def extract_excel_pure(file_content: bytes, filename: str) -> List[Dict[str, Any]]:
    """
    Extract structured data from Excel files using PURE PYTHON only.
    NO pandas, NO numpy, NO compiled dependencies.
    
    Returns:
        List of sheet dictionaries for the extractor to use.
    """
    ext = Path(filename).suffix.lower()
    
    if ext == '.xlsx':
        return _extract_xlsx(file_content)
    elif ext == '.xls':
        return _extract_xls(file_content)
    else:
        return []


def _extract_xlsx(file_content: bytes) -> List[Dict[str, Any]]:
    """Extract from .xlsx using openpyxl (pure Python)."""
    try:
        import openpyxl
        from openpyxl import load_workbook
    except ImportError:
        return []
    
    try:
        workbook = load_workbook(
            filename=io.BytesIO(file_content),
            read_only=True,
            data_only=True  # Get values, not formulas
        )
        
        all_sheets_data = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Get all rows
            rows = []
            headers = []
            
            for row in sheet.iter_rows(values_only=True):
                # Skip completely empty rows
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue
                
                # First non-empty row is headers
                if not headers:
                    headers = [str(cell).strip() if cell is not None else f"Column{i+1}" 
                              for i, cell in enumerate(row)]
                    continue
                
                # Build row data
                row_values = []
                for cell in row:
                    val = str(cell).strip() if cell is not None else ""
                    # Escape pipes for markdown
                    val = val.replace("|", "\\|")
                    row_values.append(val)
                
                if any(row_values):
                    rows.append(row_values)
            
            if rows:
                # Create proper Markdown table
                md_lines = []
                md_lines.append(f"## Sheet: {sheet_name}")
                md_lines.append("")
                md_lines.append("| " + " | ".join(headers) + " |")
                md_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
                for r in rows:
                    md_lines.append("| " + " | ".join(r) + " |")
                
                all_sheets_data.append({
                    "sheet_name": sheet_name,
                    "headers": headers,
                    "num_rows": len(rows),
                    "num_cols": len(headers),
                    "markdown": "\n".join(md_lines)
                })
        
        workbook.close()
        return all_sheets_data
        
    except Exception as e:
        print(f"Error in _extract_xlsx: {e}")
        return []


def _extract_xls(file_content: bytes) -> List[Dict[str, Any]]:
    """Extract from .xls using xlrd 1.2.0 (pure Python)."""
    try:
        import xlrd
    except ImportError:
        return []
    
    try:
        workbook = xlrd.open_workbook(file_contents=file_content)
        all_sheets_data = []
        
        for sheet_idx in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_idx)
            
            if sheet.nrows == 0:
                continue
            
            # Get headers from first row
            headers = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(0, col_idx)
                headers.append(str(cell.value).strip() if cell.value else f"Column{col_idx+1}")
            
            # Get data rows
            rows = []
            for row_idx in range(1, sheet.nrows):
                row_values = []
                has_data = False
                
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    val = str(cell.value).strip() if cell.value is not None else ""
                    # Escape pipes
                    val = val.replace("|", "\\|")
                    row_values.append(val)
                    if val:
                        has_data = True
                
                if has_data:
                    rows.append(row_values)
            
            if rows:
                # Create proper Markdown table
                md_lines = []
                md_lines.append(f"## Sheet: {sheet.name}")
                md_lines.append("")
                md_lines.append("| " + " | ".join(headers) + " |")
                md_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
                for r in rows:
                    md_lines.append("| " + " | ".join(r) + " |")
                
                all_sheets_data.append({
                    "sheet_name": sheet.name,
                    "headers": headers,
                    "num_rows": len(rows),
                    "num_cols": len(headers),
                    "markdown": "\n".join(md_lines)
                })
        
        return all_sheets_data
        
    except Exception as e:
        print(f"Error in _extract_xls: {e}")
        return []



def get_excel_metadata(file_content: bytes, filename: str) -> Dict[str, Any]:
    """
    Get basic metadata without pandas.
    
    Returns:
        Dict with sheet names and row counts
    """
    ext = Path(filename).suffix.lower()
    
    try:
        if ext == '.xlsx':
            import openpyxl
            workbook = openpyxl.load_workbook(
                filename=io.BytesIO(file_content),
                read_only=True,
                data_only=True
            )
            sheets = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                row_count = sum(1 for _ in sheet.iter_rows())
                sheets.append({"name": sheet_name, "rows": row_count})
            workbook.close()
            return {"sheets": sheets, "format": "xlsx"}
            
        elif ext == '.xls':
            import xlrd
            workbook = xlrd.open_workbook(file_contents=file_content)
            sheets = []
            for sheet_idx in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_idx)
                sheets.append({"name": sheet.name, "rows": sheet.nrows})
            return {"sheets": sheets, "format": "xls"}
            
    except Exception as e:
        return {"error": str(e), "sheets": []}
    
    return {"error": "Unsupported format", "sheets": []}
